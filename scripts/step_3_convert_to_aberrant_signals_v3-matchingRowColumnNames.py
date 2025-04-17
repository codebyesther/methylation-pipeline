import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# STEP1: Auto-detect input files
output_dir = 'output'
files_in_dir = os.listdir(output_dir)
glob20_file = next((os.path.join(output_dir, f) for f in files_in_dir if "output_glob20" in f and f.endswith(".xlsx")), None)
globmin80_file = next((os.path.join(output_dir, f) for f in files_in_dir if "output_globmin80" in f and f.endswith(".xlsx")), None)

if not glob20_file or not globmin80_file:
    raise FileNotFoundError("One or both input files not found.")

# STEP2: Load Excel files
glob20_df = pd.read_excel(glob20_file)
globmin80_df = pd.read_excel(globmin80_file)

# STEP3: Filter CGI rows only
label_col = glob20_df.columns[0]
cpg_label = "Total CpG island fragments counts for this particular spreadsheet"
filtered_glob20 = glob20_df[glob20_df[label_col].str.startswith("CGI_")].copy()
filtered_globmin80 = globmin80_df[globmin80_df[label_col].str.startswith("CGI_")].copy()

# STEP4: Align by row and column labels
filtered_glob20.set_index(label_col, inplace=True)
filtered_globmin80.set_index(label_col, inplace=True)
shared_rows = filtered_glob20.index.intersection(filtered_globmin80.index)
shared_columns = filtered_glob20.columns.intersection(filtered_globmin80.columns)
aligned_glob20 = filtered_glob20.loc[shared_rows, shared_columns].sort_index().sort_index(axis=1)
aligned_globmin80 = filtered_globmin80.loc[shared_rows, shared_columns].sort_index().sort_index(axis=1)

# STEP5: Safe division with INF handling
with pd.option_context('mode.use_inf_as_na', True):
    ratio_raw = (aligned_glob20 / aligned_globmin80.replace(0, pd.NA)) * 100000
    ratio_df = ratio_raw.fillna("INF")

# STEP6: Count and log INF values
inf_mask = ratio_df == "INF"
inf_count = inf_mask.sum().sum()
print(f"⚠️ Total 'INF' values (division by zero): {inf_count}")

# Save INF locations to CSV
inf_locations = [(row_idx, col) for row_idx, row in ratio_df.iterrows() for col in ratio_df.columns if row[col] == "INF"]
inf_log_df = pd.DataFrame(inf_locations, columns=["CGI_Region", "Sample"])
inf_log_path = os.path.join(output_dir, "inf_locations_log.csv")
inf_log_df.to_csv(inf_log_path, index=False)
print(f"📝 Logged INF locations to: {inf_log_path}")

# Save INF summary (column-wise INF counts) to CSV
inf_summary = inf_mask.sum().reset_index()
inf_summary.columns = ["Sample", "INF_Count"]
inf_summary_path = os.path.join(output_dir, "inf_summary.csv")
inf_summary.to_csv(inf_summary_path, index=False)
print(f"📊 Saved INF summary to: {inf_summary_path}")

# STEP7: Format for Excel export
ratio_df.reset_index(inplace=True)
total_glob20_row = glob20_df[glob20_df[label_col] == cpg_label].copy()
total_globmin80_row = globmin80_df[globmin80_df[label_col] == cpg_label].copy()
if not total_glob20_row.empty:
    total_glob20_row.iloc[0, 0] = "Total CpG island fragments counts for Glob20"
if not total_globmin80_row.empty:
    total_globmin80_row.iloc[0, 0] = "Total CpG island fragments counts for GlobMin80"
result_df = pd.concat([total_glob20_row, total_globmin80_row, ratio_df], ignore_index=True)

# STEP8: Export Excel with red highlight for INF
output_file = os.path.join(output_dir, "scaled_fragment_ratios_matrix.xlsx")
result_df.to_excel(output_file, index=False)

wb = load_workbook(output_file)
ws = wb.active
ws.title = "Scaled Fragment Ratios"
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in ws.iter_rows(min_row=3, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value == "INF":
            cell.fill = red_fill

wb.save(output_file)
print(f"✅ Final Excel output saved with red-highlighted INF cells: {output_file}")
