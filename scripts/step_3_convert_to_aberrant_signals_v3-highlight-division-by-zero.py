import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# STEP1: Auto-detect input files in the "output" directory
output_dir = 'output'
files_in_dir = os.listdir(output_dir)
glob20_file = next((os.path.join(output_dir, f) for f in files_in_dir if "output_glob20" in f and f.endswith(".xlsx")), None)
globmin80_file = next((os.path.join(output_dir, f) for f in files_in_dir if "output_globmin80" in f and f.endswith(".xlsx")), None)

if not glob20_file or not globmin80_file:
    raise FileNotFoundError("One or both input files ('output_glob20_*.xlsx', 'output_globmin80_*.xlsx') not found in the 'output' directory.")

# STEP2: Load the Excel files
glob20_df = pd.read_excel(glob20_file)
globmin80_df = pd.read_excel(globmin80_file)

# STEP3: Filter rows independently for each DataFrame
filter_condition_glob20 = glob20_df.iloc[:, 0].str.contains("CGI_chr") | (glob20_df.iloc[:, 0] == "Total CpG island fragments counts for this particular spreadsheet")
filtered_glob20 = glob20_df[filter_condition_glob20]

filter_condition_globmin80 = globmin80_df.iloc[:, 0].str.contains("CGI_chr") | (globmin80_df.iloc[:, 0] == "Total CpG island fragments counts for this particular spreadsheet")
filtered_globmin80 = globmin80_df[filter_condition_globmin80]

# STEP4: Align rows by their first column (ensure they have the same labels)
aligned_glob20 = filtered_glob20[filtered_glob20.iloc[:, 0].isin(filtered_globmin80.iloc[:, 0])]
aligned_globmin80 = filtered_globmin80[filtered_globmin80.iloc[:, 0].isin(filtered_glob20.iloc[:, 0])]

# Ensure the aligned DataFrames have the same shape
assert aligned_glob20.shape == aligned_globmin80.shape, "Aligned DataFrames do not have the same shape."

# STEP5: Divide the values and flag divide-by-zero
glob20_values = aligned_glob20.iloc[:, 1:].astype(float).reset_index(drop=True)
globmin80_values = aligned_globmin80.iloc[:, 1:].astype(float).reset_index(drop=True)

# Create a mask for divide-by-zero
divide_by_zero_mask = (globmin80_values == 0)

# Perform division
ratio_df = (glob20_values / globmin80_values) * 100000

# STEP6: Reassemble the DataFrame with original row labels
result_df = pd.concat([aligned_glob20.iloc[:, 0].reset_index(drop=True), ratio_df], axis=1)

# STEP7: Add total rows
total_glob20_row = glob20_df[glob20_df.iloc[:, 0] == "Total CpG island fragments counts for this particular spreadsheet"]
total_globmin80_row = globmin80_df[globmin80_df.iloc[:, 0] == "Total CpG island fragments counts for this particular spreadsheet"]
if not total_glob20_row.empty:
    total_glob20_row.iloc[0, 0] = "Total CpG island fragments counts for Glob20"
if not total_globmin80_row.empty:
    total_globmin80_row.iloc[0, 0] = "Total CpG island fragments counts for GlobMin80"
summary_rows = pd.concat([total_glob20_row, total_globmin80_row], ignore_index=True)
result_df = pd.concat([summary_rows, result_df], ignore_index=True)

# STEP8: Save without formatting first
output_file = os.path.join(output_dir, "scaled_fragment_ratios_matrix.xlsx")
result_df.to_excel(output_file, index=False)

# STEP9: Apply red highlighting to divide-by-zero cells
wb = load_workbook(output_file)
ws = wb.active
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Offset by header row + any summary rows (2 rows)
row_offset = 3  # Assuming 2 summary rows + header
for row_idx in range(divide_by_zero_mask.shape[0]):
    for col_idx in range(divide_by_zero_mask.shape[1]):
        if divide_by_zero_mask.iat[row_idx, col_idx]:
            excel_row = row_idx + row_offset
            excel_col = col_idx + 2  # +2 because first column is label, and openpyxl is 1-based
            ws.cell(row=excel_row, column=excel_col).fill = red_fill

wb.save(output_file)
print(f"Saved output to {output_file} with red highlights for divide-by-zero cells.")
